import urllib
import time
from lxml import etree
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
import pubmed
import xlwt
import openpyxl
import compare

global pmid
pmid=[]
class NcbiInfo(object):
    option = webdriver.ChromeOptions()
    option.add_argument('headless')
    browser = webdriver.Chrome(options=option)
    start_url = 'https://www.ncbi.nlm.nih.gov/pubmed/?term='
    wait = WebDriverWait(browser, 10)


    def __init__(self, keywordlist):
        self.temp = [urllib.parse.quote(i) for i in keywordlist]
        self.keyword = '%2C'.join(self.temp)
        self.title = ' AND '.join(self.temp)
        self.url = NcbiInfo.start_url + self.keyword
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36'}
        self.file = open('information.txt', 'w')
        self.status = True
        self.yearlist = []

    def click_yearandabstract(self, ):
        self.browser.get(self.url)
        try:
            perpage = self.wait.until(EC.element_to_be_clickable(
                (By.XPATH, '//ul[@class="inline_list left display_settings"]/li[3]/a/span[4]')))
            perpage.click()
            page_200 = self.wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, '#display_settings_menu_ps > fieldset > ul > li:nth-child(6) > label')))
            page_200.click()
        except TimeoutException:
            self.status = False
            pmid.append('nanana')

    def get_response(self):
        self.html = self.browser.page_source
        self.doc = etree.HTML(self.html)

    def get_info(self):
        self.art_timeanddoi = self.doc.xpath('//div[@class="resc"]/dl[@class="rprtid"]/dd/text()')
        for i in self.art_timeanddoi:
            print(i)
            pmid.append(i)





    def next_page(self):
        try:
            self.nextpage = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, '//*[@title="Next page of results"]')))
        except TimeoutException:
            self.status = False

    def main(self):
        self.click_yearandabstract()
        time.sleep(3)
        self.get_response()
        while True:
            self.get_info()
            self.next_page()
            if self.status:
                self.nextpage.click()
                self.get_response()
            else:
                break

def savepmid(pid,key):
    book1 = xlwt.Workbook()  # 新建一个excel
    sheet = book1.add_sheet('case1_sheet')
    row = 0
    for i in pid:
        i = "https://www.ncbi.nlm.nih.gov/pubmed/" + i
        sheet.write(row, 0, i)
        row = row + 1
    book1.save(key + '.xlsx')

def readkey(hanghao):
   wb = openpyxl.load_workbook('search.xlsx')    # 加载工作薄
   wb.sheetnames                                       # 获取当前所有工作表的名称， 返回一个列表
   wb.active                                           # 获
   sheet = wb['Sheet1']                      # 获取当前活动表的名称
   key=sheet.cell(row=hanghao,column=1).value

   return key

def readrows():
   wb = openpyxl.load_workbook('search.xlsx')  # 加载工作薄
   wb.sheetnames  # 获取当前所有工作表的名称， 返回一个列表
   wb.active  # 获
   sheet = wb['Sheet1']  # 获取当前活动表的名称
   rows = sheet.max_row
   return rows

if __name__ == '__main__':
    hanghao = 0
    total=[]
    while hanghao < readrows():
       hanghao = hanghao + 1
       key = readkey(hanghao)  # 读取关键词
       print(key)  # 是否读取到关键词
       a = NcbiInfo([key])
       a.main()
       total.append(key+'共计'+str(len(pmid))+'篇文献')
       savepmid(pmid,key)
       pmid=[]
       pubmed.abstarct(hanghao)
       compare.compare(key)
    for i in total:
        print(i)
