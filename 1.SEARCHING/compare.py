import re
import openpyxl
from openpyxl import Workbook
def compare(words):
  wb = openpyxl.load_workbook(words+'_result.xlsx')
  wb.sheetnames                                       # 获取当前所有工作表的名称， 返回一个列表
  wb.active                                           # 获
  sheet = wb['table']
  keyword=open('keyword.txt')
  st=keyword.readlines()
  keyword1 = open('keyword1.txt')
  st1 = keyword1.readlines()
  keyword2 = open('keyword2.txt')
  st2 = keyword2.readlines()
  row1 =1
  wb1 = Workbook()
  sheet1 = wb1.active

  while row1 < sheet.max_row:
      row1 = row1 + 1
      have = 0
      column1 = 0
      cunchu = []
      while column1 < sheet.max_column:
          column1 = column1 + 1
          key1 = sheet.cell(row=row1, column=column1).value
          print(key1)
          cunchu.append(key1)
          if column1 > 1:
            for i in st:
              i = i.strip('\n')
              if key1 != None:
               for match in re.findall(i, key1):
                  print('find '+match)
                  have = 1
                  i = str(i)
                  print(i)
                  cunchu.append(i)
            for t in st1:
               t = t.strip('\n')
               if key1 != None:
                   for match1 in re.findall(t, key1):
                       for j in st2:
                           j = j.strip('\n')
                           for match2 in re.findall(j, key1):
                               print('find ' + match1 + ' and ' + match2)
                               have = 1
                               t = str(t)
                               j = str(j)
                               cunchu.append(t + ' and ' + j)
      if have == 1:
         sheet1.append(cunchu)
  wb1.save(words+'_complete.xlsx')
