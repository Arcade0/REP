import openpyxl
# 导入工作薄
def readkey(hanghao):
   wb = openpyxl.load_workbook('search.xlsx')    # 加载工作薄
   wb.sheetnames                                       # 获取当前所有工作表的名称， 返回一个列表
   wb.active                                           # 获
   sheet = wb['Sheet1']                      # 获取当前活动表的名称
   key=sheet.cell(row=hanghao,column=1).value
   return key

def readrows():
   wb = openpyxl.load_workbook('search.xlsx')  # 加载工作薄
   wb.sheetnames  # 获取当前所有工作表的名称， 返回一个列表
   wb.active  # 获
   sheet = wb['Sheet1']  # 获取当前活动表的名称
   rows = sheet.max_row
   return rows